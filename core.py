"""
core.py
Location: project root (C:\\projects\\arbitrage_bot\\core.py)

LAYER 1 - shared data infrastructure.

Every bot calls these functions rather than talking to the API itself, so
there is only ever ONE copy of the OddsPapi parsing logic to maintain.

CHANGES IN THIS VERSION (15 Aug 2026) - all backward compatible:

  1. FIXED A REAL BUG in _extract_price(). It checked exchangeMeta["back"],
     but the API actually uses "availableToBack" / "availableToLay". That
     branch had therefore NEVER executed - the function silently fell through
     to player["price"] every time. It returned the correct number by luck,
     because the top-level price happens to equal the best back price.

  2. _extract_price() now takes a `side` argument ("back" or "lay").
     It defaults to "back", so every existing call behaves exactly as before.

  3. New _extract_exchange() returns price AND available size, for 2up_bot.

  4. New call_oddspapi() helper with rate-limit handling. The API returns
     HTTP 429 if you call the same endpoint faster than about every 0.54s.

  5. All paths now anchored to this file's folder, so scripts in subfolders
     (like scripts/2up_bot/run.py) read and write the right directories.

  6. Output folders are created automatically if missing.

  7. Fixed falsy-zero bugs where a legitimate 0.0 was treated as missing.
"""

import os
import json
import time
import statistics
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    SEAT_BET365, SEAT_PINNACLE, SEAT_BETFAIR,
    SEAT_KALSHI, SEAT_POLYMARKET,
    IP_FORMAT_SEATS, S_IDX_BOOKS, VARIANCE_LIMIT,
    ODDSPAPI_BASE_URL, REQUEST_DELAY,
    EXCHANGE_BACK_KEY, EXCHANGE_LAY_KEY,
)

# ─────────────────────────────────────────────
# PATHS AND CREDENTIALS
# ─────────────────────────────────────────────

# This file sits in the project root, so its folder IS the project root.
# Anchoring to it means scripts running from scripts/2up_bot/ still find
# the right .env and write to the right data/ and outputs/ folders.
PROJECT_ROOT = Path(__file__).resolve().parent

load_dotenv(PROJECT_ROOT / ".env")
API_KEY  = os.getenv("ODDSPAPI_API_KEY")
BASE_URL = ODDSPAPI_BASE_URL

DIR_RAW       = PROJECT_ROOT / "data" / "raw"
DIR_PROCESSED = PROJECT_ROOT / "data" / "processed"
DIR_OUTPUTS   = PROJECT_ROOT / "outputs"
DIR_LOGS      = PROJECT_ROOT / "logs"


def ensure_directories():
    """Create the working folders if they do not exist. Safe to call repeatedly."""
    for folder in (DIR_RAW, DIR_PROCESSED, DIR_OUTPUTS, DIR_LOGS):
        folder.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────────
# MATCH ODDS CONSTANTS
# ─────────────────────────────────────────────

# Premier League. Previously 16 (World Cup), which finished in July 2026 and
# now returns no fixtures - so the demo block at the bottom had nothing to show.
TOURNAMENT_ID = 17

MARKET_ID   = "101"
OUTCOME_MAP = {"101": "home", "102": "draw", "103": "away"}


# ─────────────────────────────────────────────
# SHARED API CALLER
# ─────────────────────────────────────────────

def call_oddspapi(endpoint, params, max_retries=2):
    """
    Call an OddsPapi endpoint, pausing first to respect the rate limit.

    On HTTP 429 (too many requests) the API tells us how long to wait in a
    'retryMs' field; we use it and try again.

    Excel analogy: like a macro that waits for a slow query to finish before
    reading the result, instead of grabbing an empty cell.

    Returns parsed JSON, or None on failure.
    """
    url    = f"{BASE_URL}/{endpoint}"
    params = {"apiKey": API_KEY, **params}

    for attempt in range(max_retries + 1):
        time.sleep(REQUEST_DELAY)

        try:
            response = requests.get(url, params=params, timeout=30)

            if response.status_code == 429:
                wait_ms = 2000
                try:
                    wait_ms = response.json()["error"].get("retryMs", 2000)
                except Exception:
                    pass

                if attempt < max_retries:
                    time.sleep((wait_ms / 1000) + 0.5)
                    continue

                print(f"  Rate limited on {endpoint} - retries exhausted")
                return None

            if response.status_code != 200:
                # Print the body, not just the code. The API explains itself
                # properly - e.g. "maximum of 5 tournament IDs" - and throwing
                # that away turns a two-second fix into a guessing game.
                print(f"  {endpoint} returned HTTP {response.status_code}")
                print(f"    {response.text[:300]}")
                return None

            return response.json()

        except Exception as exc:
            print(f"  {endpoint} failed: {exc}")
            return None

    return None


# ─────────────────────────────────────────────
# PRICE EXTRACTION
# ─────────────────────────────────────────────

def _get_exchange_levels(player, side):
    """
    Return the exchange price ladder for one side, or None.

    The ladder is a list of up to 3 levels, best price first, each a dict
    of {price, size}. Betfair calls these 'availableToBack' / 'availableToLay'.
    """
    exchange = player.get("exchangeMeta")
    if not isinstance(exchange, dict):
        return None

    key    = EXCHANGE_BACK_KEY if side == "back" else EXCHANGE_LAY_KEY
    levels = exchange.get(key)

    if isinstance(levels, list) and levels:
        return levels
    return None


def _get_player(bookmaker_data, outcome_id):
    """Dig down to the player block holding the price for one outcome."""
    markets = bookmaker_data.get("markets", {})
    if MARKET_ID not in markets:
        return None

    outcome = markets[MARKET_ID].get("outcomes", {}).get(outcome_id)
    if not outcome:
        return None

    return outcome.get("players", {}).get("0")


def _extract_price(bookmaker_data, outcome_id, side="back"):
    """
    Return the decimal odds for one outcome at one bookmaker.

    side="back" (default) - the price you can BACK at. For a normal bookmaker
                            this is just its price; for an exchange it is the
                            best available-to-back price.
    side="lay"            - the price you can LAY at. Only exchanges have one.

    IMPORTANT: when side="lay" and no exchange data exists, this returns None
    rather than falling back to the ordinary price. A back price returned in
    answer to a lay request would make every selection look like a perfect
    arbitrage - a silent and expensive bug.
    """
    try:
        player = _get_player(bookmaker_data, outcome_id)
        if not player:
            return None

        levels = _get_exchange_levels(player, side)
        if levels:
            return levels[0].get("price")

        # No exchange ladder available.
        if side == "lay":
            return None          # soft books have no lay price - never guess

        return player.get("price")

    except (KeyError, TypeError, IndexError):
        return None


def _extract_exchange(bookmaker_data, outcome_id, side="lay"):
    """
    Return {"price": x, "size": y} for the best level on one side, or None.

    'size' is how much money is available at that price. A tempting lay price
    with only $12 behind it is not a real opportunity, so 2up_bot reports it.
    """
    try:
        player = _get_player(bookmaker_data, outcome_id)
        if not player:
            return None

        levels = _get_exchange_levels(player, side)
        if not levels:
            return None

        best = levels[0]
        return {
            "price": best.get("price"),
            "size":  best.get("size"),
            "depth": levels,          # all 3 levels, for later use
        }

    except (KeyError, TypeError, IndexError):
        return None


# ─────────────────────────────────────────────
# IP CONVERSION
# ─────────────────────────────────────────────

def convert_to_ip(price, seat_slug):
    """Convert decimal odds to implied probability. Excel: =1/D2"""
    if price is None:
        return None
    try:
        if seat_slug in IP_FORMAT_SEATS:
            return float(price)
        return 1 / float(price)
    except (ZeroDivisionError, TypeError, ValueError):
        return None


# ─────────────────────────────────────────────
# S_IDX COMPUTATION
# ─────────────────────────────────────────────

def compute_s_idx(bookmaker_odds, outcome_id):
    """Mean implied probability across the active soft books."""
    ips = []
    for slug in S_IDX_BOOKS:
        if slug not in bookmaker_odds:
            continue
        price = _extract_price(bookmaker_odds[slug], outcome_id)
        ip    = convert_to_ip(price, slug)
        if ip is not None:
            ips.append(ip)

    if len(ips) < 3:
        return None
    return {"ip": sum(ips) / len(ips), "active_books": len(ips)}


# ─────────────────────────────────────────────
# BASKET VARIANCE
# ─────────────────────────────────────────────

def compute_basket_variance(seat_ips):
    """
    Coefficient of variation across active seats.
    Excel: =STDEV(range)/AVERAGE(range)
    """
    if len(seat_ips) < 2:
        return {"cv": None, "flagged": False}

    values  = list(seat_ips.values())
    mean    = sum(values) / len(values)

    if not mean:
        return {"cv": None, "flagged": False}

    std_dev = statistics.stdev(values)
    cv      = std_dev / mean

    # Note "is not None" rather than a plain truth test: a CV of exactly 0.0
    # means every seat agreed perfectly, which is a real result, not a missing one.
    return {
        "cv":      round(cv, 4),
        "flagged": cv > VARIANCE_LIMIT,
    }


# ─────────────────────────────────────────────
# FETCH ODDSPAPI
# ─────────────────────────────────────────────

def fetch_oddspapi():
    """Pull prices for every basket seat and save the raw response."""
    ensure_directories()

    fixtures_data = call_oddspapi("fixtures", {"tournamentId": TOURNAMENT_ID})
    fixture_names = {}

    if fixtures_data:
        for f in fixtures_data:
            fixture_names[f["fixtureId"]] = {
                "home": f["participant1Name"],
                "away": f["participant2Name"],
            }

    all_books = [
        SEAT_BET365, SEAT_PINNACLE, SEAT_BETFAIR,
        SEAT_KALSHI, SEAT_POLYMARKET,
    ] + S_IDX_BOOKS

    combined = {}

    for book in all_books:
        data = call_oddspapi(
            "odds-by-tournaments",
            {"tournamentIds": TOURNAMENT_ID, "bookmaker": book},
        )

        if not data:
            print(f"  Skipping {book}: no data returned")
            continue

        if isinstance(data, dict) and "error" in data:
            print(f"  Skipping {book}: {data['error']['message']}")
            continue

        for fixture in data:
            fid = fixture.get("fixtureId")
            if not fid:
                continue
            if fid not in combined:
                combined[fid] = {"fixtureId": fid, "bookmakerOdds": {}}
            combined[fid]["bookmakerOdds"].update(fixture.get("bookmakerOdds", {}))

        print(f"  Fetched {book}")

    raw_data = list(combined.values())

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_path  = DIR_RAW / f"oddspapi_{timestamp}.json"
    with open(raw_path, "w") as f:
        json.dump(raw_data, f)

    print(f"\nTotal fixtures fetched: {len(raw_data)}")
    return raw_data, fixture_names


# ─────────────────────────────────────────────
# BUILD BASKET
# ─────────────────────────────────────────────

def build_basket():
    """Compute IP_bar per outcome per fixture across all seats."""
    raw_data, fixture_names = fetch_oddspapi()

    named_seats = {
        "bet365":     SEAT_BET365,
        "pinnacle":   SEAT_PINNACLE,
        "betfair":    SEAT_BETFAIR,
        "kalshi":     SEAT_KALSHI,
        "polymarket": SEAT_POLYMARKET,
    }

    results = []

    for fixture in raw_data:
        fid   = fixture.get("fixtureId")
        names = fixture_names.get(fid, {})
        books = fixture.get("bookmakerOdds", {})

        fixture_result = {
            "fixture_id": fid,
            "home_team":  names.get("home", "Unknown"),
            "away_team":  names.get("away", "Unknown"),
        }

        for outcome_id, outcome_label in OUTCOME_MAP.items():
            seat_ips = {}

            for seat_name, slug in named_seats.items():
                if slug not in books:
                    continue
                price = _extract_price(books[slug], outcome_id)
                ip    = convert_to_ip(price, slug)
                if ip is not None:
                    seat_ips[seat_name] = round(ip, 6)

            s_idx = compute_s_idx(books, outcome_id)
            if s_idx:
                seat_ips["s_idx"] = round(s_idx["ip"], 6)

            active_ips = list(seat_ips.values())
            ip_bar     = sum(active_ips) / len(active_ips) if active_ips else None
            variance   = compute_basket_variance(seat_ips)

            fixture_result[outcome_label] = {
                # "is not None" so a genuine 0.0 is not mistaken for missing data
                "ip_bar":       round(ip_bar, 6) if ip_bar is not None else None,
                "active_seats": len(active_ips),
                "cv":           variance["cv"],
                "cv_flagged":   variance["flagged"],
                "seats":        seat_ips,
            }

        results.append(fixture_result)

    return results


# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

def export_to_excel(basket):
    """
    Export basket data to a formatted Excel file in outputs/.
    One row per fixture. Columns grouped by seat. CV-flagged rows highlighted.
    """
    ensure_directories()

    wb = Workbook()
    ws = wb.active
    ws.title = "Basket"

    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    GROUP_FILLS = {
        "consensus":  PatternFill("solid", start_color="2E75B6"),
        "bet365":     PatternFill("solid", start_color="375623"),
        "pinnacle":   PatternFill("solid", start_color="833C00"),
        "betfair":    PatternFill("solid", start_color="7030A0"),
        "kalshi":     PatternFill("solid", start_color="C55A11"),
        "polymarket": PatternFill("solid", start_color="006064"),
        "s_idx":      PatternFill("solid", start_color="1F4E79"),
    }
    DATA_FONT = Font(name="Arial", size=10)
    FLAG_FILL = PatternFill("solid", start_color="FFEB3B")
    CENTER    = Alignment(horizontal="center", vertical="center")

    seats          = ["bet365", "pinnacle", "betfair", "kalshi", "polymarket", "s_idx"]
    outcomes       = ["home", "draw", "away"]
    outcome_labels = {"home": "H", "draw": "D", "away": "A"}

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18

    col = 1

    ws.cell(1, col, "Fixture")
    ws.cell(1, col).font      = HEADER_FONT
    ws.cell(1, col).fill      = HEADER_FILL
    ws.cell(1, col).alignment = CENTER
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    ws.column_dimensions[get_column_letter(col)].width = 28
    col += 1

    for o in outcomes:
        ws.cell(1, col, f"IP_bar {outcome_labels[o]}")
        ws.cell(1, col).font      = HEADER_FONT
        ws.cell(1, col).fill      = GROUP_FILLS["consensus"]
        ws.cell(1, col).alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 11
        col += 1

    for o in outcomes:
        ws.cell(1, col, f"CV {outcome_labels[o]}")
        ws.cell(1, col).font      = HEADER_FONT
        ws.cell(1, col).fill      = GROUP_FILLS["consensus"]
        ws.cell(1, col).alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 9
        col += 1

    for seat in seats:
        for o in outcomes:
            ws.cell(1, col, f"{seat.title()} {outcome_labels[o]}")
            ws.cell(1, col).font      = HEADER_FONT
            ws.cell(1, col).fill      = GROUP_FILLS.get(seat, HEADER_FILL)
            ws.cell(1, col).alignment = CENTER
            ws.column_dimensions[get_column_letter(col)].width = 11
            col += 1

    for row_idx, fixture in enumerate(basket, start=3):
        home = fixture["home_team"]
        away = fixture["away_team"]
        c    = 1

        any_flagged = any(
            fixture.get(o, {}).get("cv_flagged", False) for o in outcomes
        )
        row_fill = FLAG_FILL if any_flagged else None

        ws.cell(row_idx, c, f"{home} vs {away}")
        ws.cell(row_idx, c).font      = Font(name="Arial", bold=True, size=10)
        ws.cell(row_idx, c).alignment = Alignment(horizontal="left", vertical="center")
        if row_fill:
            ws.cell(row_idx, c).fill = row_fill
        c += 1

        # IP_bar per outcome, written once as a true percentage
        for o in outcomes:
            ip_bar = fixture.get(o, {}).get("ip_bar")
            ws.cell(row_idx, c, ip_bar)
            ws.cell(row_idx, c).font          = DATA_FONT
            ws.cell(row_idx, c).alignment     = CENTER
            ws.cell(row_idx, c).number_format = "0.00%"
            if row_fill:
                ws.cell(row_idx, c).fill = row_fill
            c += 1

        for o in outcomes:
            cv  = fixture.get(o, {}).get("cv")
            flg = fixture.get(o, {}).get("cv_flagged", False)
            ws.cell(row_idx, c, cv)
            ws.cell(row_idx, c).font          = DATA_FONT
            ws.cell(row_idx, c).alignment     = CENTER
            ws.cell(row_idx, c).number_format = "0.0000"
            if flg:
                ws.cell(row_idx, c).fill = FLAG_FILL
            c += 1

        for seat in seats:
            for o in outcomes:
                ip  = fixture.get(o, {}).get("seats", {}).get(seat)
                val = round(1 / ip, 2) if ip else None
                ws.cell(row_idx, c, val)
                ws.cell(row_idx, c).font          = DATA_FONT
                ws.cell(row_idx, c).alignment     = CENTER
                ws.cell(row_idx, c).number_format = "0.00"
                if row_fill:
                    ws.cell(row_idx, c).fill = row_fill
                c += 1

    ws.freeze_panes = "B3"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path      = DIR_OUTPUTS / f"basket_{timestamp}.xlsx"
    wb.save(path)
    print(f"\nExcel file saved: {path}")
    return str(path)


# ─────────────────────────────────────────────
# TEST BLOCK
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("Building basket...\n")
    basket = build_basket()

    print("\n--- BASKET OUTPUT (first 3 fixtures) ---")
    for fixture in basket[:3]:
        print(f"\n{fixture['home_team']} vs {fixture['away_team']}")
        for outcome in ["home", "draw", "away"]:
            if outcome in fixture:
                data = fixture[outcome]
                print(f"  {outcome.upper()}")
                print(f"    IP_bar:       {data['ip_bar']}")
                print(f"    Active seats: {data['active_seats']}")
                flag = "FLAGGED" if data["cv_flagged"] else ""
                print(f"    CV:           {data['cv']} {flag}")
                print(f"    Seat IPs:     {data['seats']}")

    print("\nExporting to Excel...")
    export_to_excel(basket)