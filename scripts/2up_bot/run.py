"""
run.py
Location: scripts/2up_bot/run.py

2UP_BOT v1 - the signal engine.

WHAT IT DOES
Backs a team at bet365 (which pays out early if that team goes two goals
ahead) and lays the same team on Betfair Exchange. If the team never goes two
clear, the two bets cancel out for a small loss. If the team DOES go two
clear and then fails to win, both bets pay - and that is where the profit is.

This script finds selections where that small loss is no worse than the
threshold set by MAX_LOSS in config.py (currently 5%).

NOTE: at 5% the estimated edge is slightly NEGATIVE - see the warning above
MAX_LOSS in config.py. Treat this setting as a discovery filter for seeing
how the market is priced, not as a signal to bet everything it returns.

THE MATHS
    B = bet365 back odds
    L = Betfair Exchange best lay odds
    S = back stake

    Lay stake  = (B x S) / L
    Liability  = Lay stake x (L - 1)
    Worst case = (B / L) - 1        <- profit and loss; NEGATIVE means a loss

    Qualifies if worst case >= MAX_LOSS (-0.05, i.e. lose no more than 5%)

    A POSITIVE worst case means Betfair's lay is shorter than bet365's back:
    a locked-in profit before the promotion even matters. A true arbitrage.

Commission is deliberately excluded - rates vary by user and Smarkets is free.

ONLY the two team selections are evaluated. The draw can never trigger 2Up
(no team to go two goals ahead), so it is never read.

API BUDGET (confirmed against the live API 15 Aug 2026)
  - /odds-by-tournaments takes at most 5 tournament IDs per call,
    so 7 leagues = 2 calls per bookmaker = 4 calls.
  - The odds response already carries startTime and tournamentId, so we
    filter to the date window BEFORE looking anything else up.
  - It does NOT carry team names, only participant IDs. Names come from
    /fixtures, which takes one tournamentId at a time - but accepts
    sportId + from/to instead, letting us get them all in ONE call.

  Best case: 5 requests per run. Worst case: 4 + one per active league.
  The bot reports which path it used.
"""

import sys
import csv
from datetime import datetime, timedelta, timezone
from pathlib import Path

# run.py lives two folders below the project root, so we add the root to
# Python's search path in order to import core.py and config.py from there.
PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

import core
import config

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


LEAGUE_NAMES = {
    17:  "Premier League (ENG)",
    18:  "Championship (ENG)",
    24:  "League One (ENG)",
    25:  "League Two (ENG)",
    36:  "Premiership (SCO)",
    8:   "LaLiga (ESP)",
    35:  "Bundesliga (GER)",
    23:  "Serie A (ITA)",
    34:  "Ligue 1 (FRA)",
    325: "Brasileiro Serie A (BRA)",
}


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def batches(items, size):
    """Split a list into chunks of at most `size`. Excel: like paging a range."""
    for start in range(0, len(items), size):
        yield items[start:start + size]


def date_chunks(start, end, max_days):
    """
    Split a date range into pieces no longer than max_days.

    The fixtures endpoint refuses a sportId query spanning 10 days or more,
    so a 21-day window becomes three 9-day chunks.
    """
    current = start
    while current < end:
        stop = min(current + timedelta(days=max_days), end)
        yield current, stop
        current = stop


def parse_kickoff(value):
    """Turn an API timestamp into a timezone-aware datetime, or None."""
    if value is None:
        return None

    if isinstance(value, (int, float)):
        try:
            seconds = value / 1000 if value > 1e11 else value
            return datetime.fromtimestamp(seconds, tz=timezone.utc)
        except (ValueError, OSError):
            return None

    try:
        text = str(value).replace("Z", "+00:00")
        parsed = datetime.fromisoformat(text)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed
    except ValueError:
        return None


# ─────────────────────────────────────────────
# STEP 1 - ODDS (also gives us times and leagues)
# ─────────────────────────────────────────────

def fetch_book_odds(tournament_ids, bookmaker):
    """
    Get all odds for one bookmaker across our leagues.

    Batched to MAX_TOURNAMENTS_PER_REQUEST because the API rejects more than
    five tournament IDs in one call.

    Returns (odds_by_fixture, meta_by_fixture, requests_used).
    """
    odds       = {}
    meta       = {}
    used       = 0
    batch_size = config.max_tournaments_for(bookmaker)
    seen       = set()

    for chunk in batches(tournament_ids, batch_size):
        id_list = ",".join(str(t) for t in chunk)
        data = core.call_oddspapi(
            "odds-by-tournaments",
            {"tournamentIds": id_list, "bookmaker": bookmaker},
        )
        used += 1

        if data is None:
            # A failed batch silently drops every league in it. Say so loudly -
            # this exact failure cost us five leagues before it was noticed.
            print(f"  WARNING: batch [{id_list}] returned nothing - "
                  f"{len(chunk)} league(s) missing from this bookmaker")
            continue

        for fixture in data:
            fid = fixture.get("fixtureId")
            if not fid:
                continue

            books = fixture.get("bookmakerOdds", {})
            if bookmaker in books:
                odds[fid] = books[bookmaker]

            tid = fixture.get("tournamentId")
            if tid:
                seen.add(tid)

            # The odds response carries these for free - no extra call needed.
            meta[fid] = {
                "kickoff":    parse_kickoff(fixture.get("startTime")),
                "tournament": tid,
                "p1_id":      fixture.get("participant1Id"),
                "p2_id":      fixture.get("participant2Id"),
            }

    missing = [t for t in tournament_ids if t not in seen]
    if missing:
        names = ", ".join(LEAGUE_NAMES.get(t, str(t)) for t in missing)
        print(f"  Note: no fixtures returned for {names}")

    return odds, meta, used


# ─────────────────────────────────────────────
# STEP 2 - TEAM NAMES (the only thing odds lack)
# ─────────────────────────────────────────────

def fetch_team_names(needed_ids, tournament_ids, window_start, window_end):
    """
    Look up team names for the fixtures we actually care about.

    Tries ONE bulk call using sportId + from/to dates. The API's own error
    message confirmed those parameters exist:
        "At least one of 'tournamentId', 'sportId', 'participantId',
         'from' or 'to' must be provided."

    Falls back to one call per league if that does not work.

    Returns (names_by_fixture, requests_used, note).
    """
    names = {}
    used  = 0

    def absorb(records):
        for f in records or []:
            fid = f.get("fixtureId")
            if fid and fid in needed_ids:
                names[fid] = (
                    f.get("participant1Name", "?"),
                    f.get("participant2Name", "?"),
                )

    # Attempt 1 - bulk by sport, split into chunks because the API rejects a
    # sportId query spanning 10 days or more.
    chunks = list(date_chunks(window_start, window_end,
                              config.FIXTURES_MAX_DATE_RANGE_DAYS))
    for chunk_start, chunk_end in chunks:
        data = core.call_oddspapi("fixtures", {
            "sportId": config.SPORT_ID_SOCCER,
            "from":    chunk_start.strftime("%Y-%m-%d"),
            "to":      chunk_end.strftime("%Y-%m-%d"),
        })
        used += 1
        absorb(data)
        if len(names) >= len(needed_ids):
            break

    if len(names) >= len(needed_ids):
        return names, used, f"bulk by sport ({used} date chunks)"

    # Attempt 2 - one call per league, but only leagues we still need.
    missing_leagues = sorted(tournament_ids)
    for tid in missing_leagues:
        data = core.call_oddspapi("fixtures", {"tournamentId": tid})
        used += 1
        absorb(data)
        if len(names) >= len(needed_ids):
            break

    got = len(names)
    return names, used, f"per-league fallback ({got}/{len(needed_ids)} named)"


# ─────────────────────────────────────────────
# THE CALCULATION
# ─────────────────────────────────────────────

def evaluate_selection(back_odds, lay_price, lay_size, stake):
    """
    Work out the numbers for one team selection.
    Excel equivalent: =(B2/L2)-1 for the worst case, =(B2*$S$1)/L2 for lay stake.
    """
    if not back_odds or not lay_price or lay_price <= 1:
        return None

    worst_case = (back_odds / lay_price) - 1
    lay_stake  = (back_odds * stake) / lay_price
    liability  = lay_stake * (lay_price - 1)

    return {
        "back_odds":  round(back_odds, 3),
        "lay_price":  round(lay_price, 3),
        "lay_size":   round(lay_size, 2) if lay_size is not None else None,
        "worst_case": worst_case,
        "back_stake": round(stake, 2),
        "lay_stake":  round(lay_stake, 2),
        "liability":  round(liability, 2),
    }


def scan():
    """Run the full scan and return (signals, stats)."""
    tournament_ids = config.TWO_UP_TOURNAMENT_IDS
    now            = datetime.now(timezone.utc)
    window_end     = now + timedelta(days=config.FIXTURE_WINDOW_DAYS)

    n_batches = -(-len(tournament_ids) // config.MAX_TOURNAMENTS_PER_REQUEST)
    print(f"Fetching {config.TWO_UP_SOFT_BOOK} odds "
          f"({len(tournament_ids)} leagues in {n_batches} batches) ...")
    soft_odds, soft_meta, req_soft = fetch_book_odds(
        tournament_ids, config.TWO_UP_SOFT_BOOK)
    print(f"  {len(soft_odds)} fixtures priced")

    print(f"Fetching {config.TWO_UP_EXCHANGE} odds ...")
    exch_odds, exch_meta, req_exch = fetch_book_odds(
        tournament_ids, config.TWO_UP_EXCHANGE)
    print(f"  {len(exch_odds)} fixtures priced")

    meta = {**exch_meta, **soft_meta}

    # Narrow to fixtures that are in the window AND priced by both books,
    # BEFORE spending any requests on team names.
    candidates = []
    for fid in soft_odds:
        if fid not in exch_odds:
            continue
        kickoff = meta.get(fid, {}).get("kickoff")
        if kickoff is not None and not (now <= kickoff <= window_end):
            continue
        candidates.append(fid)

    print(f"In window and priced by both: {len(candidates)} fixtures")

    names, req_names, note = ({}, 0, "not needed")
    if candidates:
        print("Fetching team names ...")
        active_leagues = sorted({
            meta[f]["tournament"] for f in candidates
            if meta.get(f, {}).get("tournament")
        })
        names, req_names, note = fetch_team_names(
            set(candidates), active_leagues or tournament_ids, now, window_end)
        print(f"  {len(names)} named  [{note}]")

    signals = []
    stats   = {"candidates": len(candidates), "evaluated": 0,
               "no_lay": 0, "too_small": 0, "qualified": 0}

    outcome_labels = {config.OUTCOME_HOME: "Home", config.OUTCOME_AWAY: "Away"}

    for fid in candidates:
        info    = meta.get(fid, {})
        kickoff = info.get("kickoff")
        home, away = names.get(fid, (f"Team {info.get('p1_id')}",
                                     f"Team {info.get('p2_id')}"))

        for outcome_id in config.TWO_UP_OUTCOMES:
            stats["evaluated"] += 1

            back_odds = core._extract_price(soft_odds[fid], outcome_id, "back")
            lay       = core._extract_exchange(exch_odds[fid], outcome_id, "lay")

            if not back_odds or not lay or not lay.get("price"):
                stats["no_lay"] += 1
                continue

            result = evaluate_selection(
                back_odds, lay["price"], lay.get("size"), config.BACK_STAKE)
            if not result or result["worst_case"] < config.MAX_LOSS:
                continue

            if config.MIN_LAY_SIZE > 0:
                size = result["lay_size"]
                if size is None or size < config.MIN_LAY_SIZE:
                    stats["too_small"] += 1
                    continue

            team = home if outcome_id == config.OUTCOME_HOME else away

            signals.append({
                "fixture":    f"{home} vs {away}",
                "kickoff":    kickoff,
                "league":     LEAGUE_NAMES.get(info.get("tournament"),
                                               str(info.get("tournament"))),
                "selection":  outcome_labels[outcome_id],
                "team":       team,
                "fixture_id": fid,
                **result,
            })
            stats["qualified"] += 1

    # Best first: a positive worst case (true arb) sorts to the top.
    signals.sort(key=lambda s: s["worst_case"], reverse=True)

    stats["total_requests"] = req_soft + req_exch + req_names
    return signals, stats


# ─────────────────────────────────────────────
# OUTPUT
# ─────────────────────────────────────────────

COLUMNS = [
    ("Fixture",      34), ("Kick-off",    17), ("League",      22),
    ("Sel",           6), ("Team",        22), ("bet365 (B)",  11),
    ("BFX Lay (L)",  11), ("Lay Size",    11), ("Worst Case",  12),
    ("Back Stake",   11), ("Lay Stake",   11), ("Liability",   11),
]


def write_excel(signals):
    """Write qualifying signals to a timestamped spreadsheet."""
    core.ensure_directories()

    wb = Workbook()
    ws = wb.active
    ws.title = "2Up Signals"

    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    DATA_FONT   = Font(name="Arial", size=10)
    ARB_FILL    = PatternFill("solid", start_color="C6EFCE")   # green = true arb
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left", vertical="center")

    for i, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, i, title)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = width

    for row, s in enumerate(signals, start=2):
        is_arb = s["worst_case"] > 0
        ko     = s["kickoff"].strftime("%d/%m/%Y %H:%M") if s["kickoff"] else "?"

        values = [
            s["fixture"], ko, s["league"], s["selection"], s["team"],
            s["back_odds"], s["lay_price"], s["lay_size"], s["worst_case"],
            s["back_stake"], s["lay_stake"], s["liability"],
        ]

        for i, value in enumerate(values, start=1):
            cell = ws.cell(row, i, value)
            cell.font      = DATA_FONT
            cell.alignment = LEFT if i in (1, 3, 5) else CENTER
            if i in (6, 7):
                cell.number_format = "0.00"
            elif i == 9:
                cell.number_format = "0.00%"
            elif i in (8, 10, 11, 12):
                cell.number_format = "#,##0.00"
            if is_arb:
                cell.fill = ARB_FILL

    ws.freeze_panes = "A2"

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path  = core.DIR_OUTPUTS / f"2up_{stamp}.xlsx"
    wb.save(path)
    return path


def append_csv(signals):
    """Append qualifying signals to the running log. Signals only."""
    core.ensure_directories()
    path      = core.DIR_LOGS / "2up_signals.csv"
    is_new    = not path.exists()
    timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")

    with open(path, "a", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        if is_new:
            writer.writerow([
                "logged_at", "fixture_id", "fixture", "kickoff", "league",
                "selection", "team", "back_odds", "lay_price", "lay_size",
                "worst_case", "back_stake", "lay_stake", "liability",
            ])
        for s in signals:
            writer.writerow([
                timestamp, s["fixture_id"], s["fixture"],
                s["kickoff"].isoformat() if s["kickoff"] else "",
                s["league"], s["selection"], s["team"],
                s["back_odds"], s["lay_price"], s["lay_size"],
                round(s["worst_case"], 5), s["back_stake"],
                s["lay_stake"], s["liability"],
            ])
    return path


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print("\n" + "=" * 78)
    print("  2UP_BOT v1  -  bet365 back  vs  Betfair Exchange lay")
    print("=" * 78)
    print(f"  Leagues:     {len(config.TWO_UP_TOURNAMENT_IDS)}")
    print(f"  Window:      next {config.FIXTURE_WINDOW_DAYS} days")
    print(f"  Back stake:  ${config.BACK_STAKE:,.2f}")
    print(f"  Max loss:    {config.MAX_LOSS * 100:.1f}%  (commission excluded)")
    if config.MIN_LAY_SIZE > 0:
        print(f"  Min lay size: ${config.MIN_LAY_SIZE:,.2f}")
    print("=" * 78 + "\n")

    signals, stats = scan()

    print("\n" + "-" * 78)
    print(f"  Candidate fixtures: {stats['candidates']}"
          f"  |  Selections checked: {stats['evaluated']}")
    print(f"  No lay price: {stats['no_lay']}"
          f"  |  Below min size: {stats['too_small']}"
          f"  |  QUALIFYING: {stats['qualified']}")
    print(f"  API requests used: {stats['total_requests']}")
    print("-" * 78)

    if not signals:
        print("\n  No qualifying selections this run.\n")
        return

    print(f"\n  {'Team':<22} {'B':>6} {'L':>6} {'Worst':>8} {'Lay $':>10}  Fixture")
    print("  " + "-" * 74)
    for s in signals[:15]:
        print(f"  {s['team'][:22]:<22} {s['back_odds']:>6.2f} {s['lay_price']:>6.2f} "
              f"{s['worst_case'] * 100:>7.2f}% {s['lay_stake']:>10,.2f}  {s['fixture'][:28]}")

    if len(signals) > 15:
        print(f"  ... and {len(signals) - 15} more in the spreadsheet")

    excel_path = write_excel(signals)
    csv_path   = append_csv(signals)

    print(f"\n  Spreadsheet: {excel_path}")
    print(f"  Signal log:  {csv_path}\n")


if __name__ == "__main__":
    main()