# scripts/freeze_bot/build_freeze_sheet.py
"""
Builds the formatted Excel workbook from a freeze screener CSV.

Sheet 1 "Candidates" — selections clearing both gates, sorted by % difference
                       (biggest gap first).
Sheet 2 "All scanned" — every selection at MIN_MATCH_ODDS or longer, including
                        failures with their reason, for later tuning.

Also refreshes output/freeze/latest.xlsx — a fixed-name copy of the newest
workbook, so you never hunt for the current one. The timestamped archive
still builds up alongside it.

Costs ZERO API requests — it only reads a CSV the runner already wrote.

Usage:
    python scripts/freeze_bot/build_freeze_sheet.py
    python scripts/freeze_bot/build_freeze_sheet.py output/freeze/freeze_candidates_20260818_234820.csv
"""

from __future__ import annotations

import csv
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

import config_freeze as cfg  # noqa: E402

# Styling
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTRE = Alignment(horizontal="center", vertical="center")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# UK short date with kick-off time.
DATE_FORMAT = "dd/mm/yyyy hh:mm"
DATE_DISPLAY_WIDTH = len("dd/mm/yyyy hh:mm")

# Columns are (csv_key, header, number_format). Widths are measured from the
# data at build time rather than fixed, so nothing is clipped or padded.
CANDIDATE_COLUMNS = [
    ("fixture", "Fixture", None),
    ("league", "League", None),
    ("kickoff_dt", "Date", DATE_FORMAT),
    ("team", "Freeze Target", None),
    ("match_odds_proxy", "Match Odds (PP)", "0.00"),
    ("ftts_odds", "FTTS Odds", "0.00"),
    ("pct_diff", "% Difference", '0.0"%"'),
]

ALL_COLUMNS = CANDIDATE_COLUMNS + [
    ("ratio", "Ratio", "0.00"),
    ("p_ftts", "P(Scores First)", "0.0%"),
    ("verdict", "Verdict", None),
    ("fail_reason", "Fail Reason", None),
]

# Text and dates read better left-aligned; figures line up centred.
LEFT_KEYS = {"fixture", "league", "kickoff_dt", "team", "verdict", "fail_reason"}

# Excel reserves roughly three characters for the autofilter dropdown arrow.
FILTER_ARROW_PAD = 3
CELL_PAD = 1
MAX_WIDTH = 44


def output_dir() -> Path:
    return PROJECT_ROOT / cfg.OUTPUT_DIR


def newest_csv() -> Path:
    out_dir = output_dir()
    if not out_dir.exists():
        sys.exit(f"ERROR: no output folder at {out_dir}")

    files = sorted(out_dir.glob("freeze_candidates_*.csv"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        sys.exit(f"ERROR: no freeze_candidates_*.csv in {out_dir}\n"
                 "       Run run_freeze.py first.")
    return files[0]


def read_rows(path: Path) -> list:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def as_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_kickoff(raw: str):
    """
    Convert the feed's UTC timestamp to a naive local datetime.

    The feed returns UTC like '2026-08-21T19:00:00.000Z'. astimezone() with
    no argument converts to this machine's timezone, so BST is handled by
    Windows rather than a hardcoded offset. Excel cannot store a timezone,
    so tzinfo is stripped after conversion.
    """
    if not raw:
        return None
    text = str(raw).replace("Z", "+00:00")
    try:
        stamp = datetime.fromisoformat(text)
    except ValueError:
        return None
    return stamp.astimezone().replace(tzinfo=None)


def enrich(rows: list) -> list:
    """Add the derived columns the sheet needs."""
    enriched = []
    for row in rows:
        side = row.get("side", "")
        team = row.get("team", "?")
        opponent = row.get("opponent", "?")

        # 'team' is the qualifying side, which may be home or away.
        # Rebuild the fixture in true home-v-away order.
        fixture = (f"{team} v {opponent}" if side == "home"
                   else f"{opponent} v {team}")

        record = dict(row)
        record["fixture"] = fixture
        record["league"] = row.get("tournament", "")
        record["kickoff_dt"] = parse_kickoff(row.get("kickoff", ""))
        record["_pct"] = as_float(row.get("pct_diff"))
        enriched.append(record)

    return enriched


def cell_value(record: dict, key: str):
    """Return the value to write, typed so Excel sorts and filters it properly."""
    raw = record.get(key, "")

    if key == "kickoff_dt":
        return raw if isinstance(raw, datetime) else None
    if key in LEFT_KEYS:
        return raw

    numeric = as_float(raw)
    return numeric if numeric is not None else raw


def display_width(value, number_format) -> int:
    """
    Approximate how wide a cell renders, so columns can be sized to content.

    Like measuring the longest entry in a column before double-clicking the
    border in Excel — same idea, done in code.
    """
    if value is None or value == "":
        return 0
    if isinstance(value, datetime):
        return DATE_DISPLAY_WIDTH
    if isinstance(value, (int, float)):
        if number_format == "0.00":
            return len(f"{value:.2f}")
        if number_format == "0.0%":
            return len(f"{value * 100:.1f}%")
        if number_format and "%" in number_format:
            return len(f"{value:.1f}%")
        return len(str(value))
    return len(str(value))


def write_sheet(worksheet, rows: list, columns: list) -> None:
    header_row = 1

    for index, (_, label, _) in enumerate(columns, start=1):
        cell = worksheet.cell(row=header_row, column=index, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER
    worksheet.row_dimensions[header_row].height = 28

    for offset, record in enumerate(rows):
        excel_row = header_row + 1 + offset

        for index, (key, _, number_format) in enumerate(columns, start=1):
            cell = worksheet.cell(row=excel_row, column=index,
                                  value=cell_value(record, key))
            cell.border = BORDER
            cell.alignment = ALIGN_LEFT if key in LEFT_KEYS else ALIGN_CENTRE
            if number_format:
                cell.number_format = number_format
            if key == "pct_diff":
                cell.font = Font(bold=True)

    # Size each column to its widest entry.
    for index, (key, label, number_format) in enumerate(columns, start=1):
        widest = max(
            [display_width(cell_value(r, key), number_format) for r in rows]
            or [0]
        )
        needed = max(len(label) + FILTER_ARROW_PAD, widest + CELL_PAD)
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            needed, MAX_WIDTH)

    last_row = header_row + len(rows)
    if rows:
        pct_index = next(i for i, (k, _, _) in enumerate(columns, start=1)
                         if k == "pct_diff")
        pct_letter = get_column_letter(pct_index)
        worksheet.conditional_formatting.add(
            f"{pct_letter}{header_row + 1}:{pct_letter}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="63BE7B",
            ),
        )
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(columns))}{last_row}"
        )

    worksheet.freeze_panes = f"A{header_row + 1}"


def refresh_latest(source: Path) -> Path | None:
    """
    Copy the new workbook over output/freeze/latest.xlsx.

    Excel locks a file while it is open, so this can fail with a permission
    error. That is not worth failing the run over — the timestamped workbook
    is already saved — so it warns and carries on.
    """
    target = output_dir() / cfg.LATEST_WORKBOOK
    if source.resolve() == target.resolve():
        return target

    try:
        shutil.copy2(source, target)
        return target
    except PermissionError:
        print(f"  NOTE: could not update {cfg.LATEST_WORKBOOK} — it is open "
              f"in Excel. Close it and re-run this script (free), or just "
              f"open the timestamped file.")
        return None
    except OSError as error:
        print(f"  NOTE: could not update {cfg.LATEST_WORKBOOK} ({error})")
        return None


def build(csv_path: Path) -> Path:
    """Build the workbook and return its path. Importable by the runner."""
    rows = enrich(read_rows(csv_path))
    if not rows:
        raise ValueError(f"{csv_path.name} has no rows.")

    # Sort by % difference, biggest gap first. Rows with no FTTS price sort last.
    rows.sort(key=lambda r: (r["_pct"] is None, -(r["_pct"] or 0)))
    passing = [r for r in rows if r.get("verdict") == "PASS"]

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Candidates"
    write_sheet(sheet, passing, CANDIDATE_COLUMNS)

    all_sheet = workbook.create_sheet("All scanned")
    write_sheet(all_sheet, rows, ALL_COLUMNS)

    out_path = csv_path.with_suffix(".xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)

    print(f"  Candidates sheet: {len(passing)} rows")
    print(f"  All scanned:      {len(rows)} rows")
    print(f"  Written to:       {out_path}")

    return out_path


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else newest_csv()
    if not source.is_absolute():
        source = PROJECT_ROOT / source
    if not source.exists():
        sys.exit(f"ERROR: file not found: {source}")

    print(f"Reading {source.name}...")
    build(source)
    print("  API requests used: 0")


if __name__ == "__main__":
    main()